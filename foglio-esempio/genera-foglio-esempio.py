"""Genera il foglio di esempio del Radar delle opportunità (3 schede)."""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

NAVY = "1C2E6E"
DESTINAZIONE = Path(
    r"D:\Repo\Lavoro\Partecipa\Partecipa-frontend\docs\strumenti\radar-opportunita"
    r"\foglio-esempio\radar-opportunita-esempio.xlsx"
)

SCHEDE = {
    "FONTI": {
        "teste": ["Nome fonte", "Indirizzo", "Attiva?", "Note", "Destinatari del riepilogo"],
        "larghezze": [34, 46, 10, 40, 34],
        "righe": [
            ["Giovani2030 — bandi e opportunità", "https://giovani2030.it/bandi-e-opportunita/",
             "sì", "Dipartimento Politiche Giovanili: l'elenco più ricco, con le scadenze", "referente@ilmiocomune.it"],
            ["Politiche Giovanili — avvisi e bandi", "https://www.politichegiovanili.gov.it/comunicazione/avvisi-e-bandi/",
             "sì", "avvisi nazionali, servizio civile", ""],
            ["Agenzia Italiana per la Gioventù", "https://www.agenziagiovani.it/",
             "sì", "Erasmus+ e Corpo Europeo di Solidarietà", ""],
            ["Eurodesk Italy — notizie", "https://www.eurodesk.it/notizie",
             "no", "mobilità europea: attivala quando ti serve", ""],
            ["Avvisi del mio Comune", "",
             "no", "il pezzo più locale e più utile: metti qui l'indirizzo della pagina avvisi", ""],
        ],
    },
    "OPPORTUNITÀ": {
        "teste": ["Trovato il", "Fonte", "Titolo", "Chi può partecipare", "Cosa serve",
                  "Entro quando", "Link", "Quanto vale", "Stato"],
        "larghezze": [16, 24, 44, 34, 34, 14, 38, 20, 16],
        "righe": [
            ["(riga di esempio: cancellala)", "Portale Giovani Regione Campania",
             "Contributi per progetti di associazioni giovanili",
             "associazioni con soci fra i 18 e i 35 anni",
             "statuto, progetto, preventivo di spesa",
             "30/09/2026", "https://esempio.it/bando-esempio", "fino a 5.000 €", "da verificare"],
        ],
    },
    "LOG": {
        "teste": ["Quando", "Fase", "Riferimento", "Che cosa è successo"],
        "larghezze": [18, 14, 40, 70],
        "righe": [],
    },
}


def main() -> None:
    wb = Workbook()
    wb.remove(wb.active)

    for nome, dati in SCHEDE.items():
        ws = wb.create_sheet(nome)
        ws.append(dati["teste"])

        for cella in ws[1]:
            cella.font = Font(bold=True, color="FFFFFF")
            cella.fill = PatternFill("solid", fgColor=NAVY)
            cella.alignment = Alignment(vertical="center", wrap_text=True)

        ws.row_dimensions[1].height = 30
        ws.freeze_panes = "A2"

        for i, larghezza in enumerate(dati["larghezze"], start=1):
            ws.column_dimensions[get_column_letter(i)].width = larghezza

        for riga in dati["righe"]:
            ws.append(riga)
            for cella in ws[ws.max_row]:
                cella.alignment = Alignment(vertical="top", wrap_text=True)

        if nome == "FONTI":
            regola = DataValidation(
                type="list", formula1='"sì,no"', allow_blank=True, showDropDown=False,
                error="Solo «sì» o «no»: le fonti con «no» vengono saltate.",
                errorTitle="Valore non valido",
            )
            ws.add_data_validation(regola)
            regola.add("C2:C200")

    DESTINAZIONE.parent.mkdir(parents=True, exist_ok=True)
    wb.save(DESTINAZIONE)
    print(f"scritto: {DESTINAZIONE} ({DESTINAZIONE.stat().st_size} byte)")


if __name__ == "__main__":
    main()
